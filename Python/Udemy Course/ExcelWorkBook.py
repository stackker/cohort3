from openpyxl import Workbook


wb = Workbook()
wsheet = wb.active
ws1 = wb.create_sheet('MySheet', 1)  # next sheet if it is 1
ws1.title = "New Title"
ws1.sheet_properties.tabColor = "1072BA"
wsheet.title = 'Red Title'

ws3 = wb['New Title']  # The ws1 worksheet is now avlb as ws3
print(wb.sheetnames)

wsheet['A4'] = 4
wsheet['A1'] = wsheet.title
wsheet.cell(row=3, column=1) 
# wsheet['B3'] = cellValue
print(wsheet.cell(row=3, column=2))
for x in range(1, 101):
    for y in range(1, 101):
        wsheet.cell(row=x, column=y, value = str(x)+'/'+str(y))

wb.save('./balances.xlsx')